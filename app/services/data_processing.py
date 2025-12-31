# -*- coding: utf-8 -*-
import pandas as pd
import unicodedata
import re
from openpyxl import load_workbook
import io
import os

# ========= CONFIGURACIÓN (Defaults) =========
PESO_POR_DEFECTO_GR = 30
ALTO_DEF = 3
ANCHO_DEF = 3
PROF_DEF = 3
VALOR_DECLARADO = 6000

HOJA_DOMICILIO = "A domicilio"
HOJA_SUCURSAL  = "A sucursal"
HOJA_CONFIG    = "Configuracion"
FILA_INICIO = 3
ULTIMA_FILA = 400

# ========= FUNCIONES AUXILIARES =========

def split_nombre_apellido(nombre_completo: str):
    if not isinstance(nombre_completo, str) or not nombre_completo.strip():
        return "", ""
    partes = nombre_completo.strip().split()
    if len(partes) == 1:
        return partes[0], ""
    return partes[0], " ".join(partes[1:])

def limpiar_numero_calle(valor):
    if not isinstance(valor, str):
        valor = str(valor)
    val = valor.strip().upper()
    if val in ("SN", "S/N", "S N", "-", ""):
        return "0"
    dig = "".join(ch for ch in val if ch.isdigit())
    if dig == "":
        return "0"
    return dig

def limpiar_telefono(raw: str):
    if not isinstance(raw, str):
        return "", ""
    txt = raw.strip()
    if not txt or "no informado" in txt.lower():
        return "", ""
    digits = "".join(ch for ch in txt if ch.isdigit())
    if not digits:
        return "", ""
    codigo = "54"
    if digits.startswith("54"):
        numero = digits[2:]
    else:
        numero = digits
    if numero.startswith("0") and len(numero) > 8:
        numero = numero[1:]
    return codigo, numero

def formatear_id(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def normalizar_texto(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    s = s.upper()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    s = s.replace(".", " ").replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def sanitizar_texto(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val)
    s = s.replace('"', '').replace("'", "")
    s = s.replace('\n', ' ').replace('\r', ' ')
    return s.strip()

def sanitizar_piso(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    # Remove invalid characters specifically mentioned by user: '-' and '/'
    # Also good to remove other common separators if they appear, but user specifically asked for these.
    # We will replace them with empty string if they are standalone or remove them from string?
    # User said: "Eliminar caracteres especificos"
    # Example: "1-A" -> "1A"
    
    # Replace bad chars with empty string
    for char in ['-', '/']:
        s = s.replace(char, '')
        
    return s.strip()

# ---------- Índices desde hoja Configuracion ----------

def construir_indice_sucursales(ws_conf):
    indice = []
    max_row = ws_conf.max_row
    # Lista simple de nombres para el frontend
    nombres_simples = []
    
    for row in range(2, max_row + 1):
        name = ws_conf.cell(row=row, column=1).value
        if not name:
            continue
        addr = ws_conf.cell(row=row, column=5).value
        name_norm = normalizar_texto(name)
        addr_norm = normalizar_texto(addr) if addr else ""
        
        # Extract province from address (Assuming format PROV / LOC / ADDR or similar)
        # Re-using logic from localidades
        prov_norm_idx = ""
        if addr:
            parts = [p.strip() for p in str(addr).split('/')]
            if len(parts) >= 1:
                prov_norm_idx = normalizar_texto(parts[0])

        indice.append((name, name_norm, addr_norm, prov_norm_idx))
        # Structure: value, context(=province/address normalized)
        nombres_simples.append({"value": name, "context": addr_norm})
        
    return indice, nombres_simples

def construir_indice_localidades(ws_conf):
    indice = []
    max_row = ws_conf.max_row
    nombres_simples = []
    
    for row in range(2, max_row + 1):
        s = ws_conf.cell(row=row, column=5).value
        if not s:
            continue
        original = str(s).strip()
        norm = normalizar_texto(original)
        parts = [p.strip() for p in str(s).split('/')]
        cp = ""
        if parts:
            last = parts[-1]
            cp_digits = "".join(ch for ch in last if ch.isdigit())
            cp = cp_digits
        loc_norm = ""
        if len(parts) >= 2:
            loc_norm = normalizar_texto(parts[1])
        prov_norm = ""
        if len(parts) >= 1:
            prov_norm = normalizar_texto(parts[0])
        indice.append((original, norm, prov_norm, loc_norm, cp))
        # Structure: value, provincia(=normalized province)
        nombres_simples.append({"value": original, "provincia": prov_norm})
        
    return indice, nombres_simples

# ---------- Matching ----------

def buscar_sucursal_por_direccion(indice_sucursales, calle, numero, localidad=None, ciudad=None, provincia=None):
    if not isinstance(calle, str) or not calle.strip():
        return "", []
    calle_norm = normalizar_texto(calle)
    tokens = calle_norm.split()
    if not tokens:
        return "", []

    # --- FILTRO POR PROVINCIA (ROBUST PROVINCE CHECK) ---
    candidatos_validos = []
    
    # Normalize input province
    prov_input_norm = ""
    if isinstance(provincia, str) and provincia.strip():
        prov_input_norm = normalizar_texto(provincia)
    
    # Handle CABA/Capital special case
    is_caba_input = "CAPITAL FEDERAL" in prov_input_norm or "CABA" in prov_input_norm or "AUTONOMA" in prov_input_norm

    for item in indice_sucursales:
        # item is now (name, name_norm, addr_norm, prov_norm_idx)
        p_idx = item[3]
        
        # If we have an input province, we MUST match it roughly
        if prov_input_norm:
            # Check CABA special match
            is_caba_idx = "CAPITAL FEDERAL" in p_idx or "CABA" in p_idx or "AUTONOMA" in p_idx
            
            if is_caba_input and is_caba_idx:
                candidatos_validos.append(item)
                continue
            elif is_caba_input != is_caba_idx:
                # One is CABA, the other is not -> Skip
                continue
                
            # General substring match (e.g. "SANTA FE" in "SANTA FE")
            # Or "BUENOS AIRES" in "PROVINCIA DE BUENOS AIRES"
            if prov_input_norm in p_idx or p_idx in prov_input_norm:
                candidatos_validos.append(item)
            else:
                # Mismatch province -> Skip
                continue
        else:
            # No input province -> Consider all
            candidatos_validos.append(item)

    # --- STREET MATCHING ON VALID CANDIDATES ONLY ---
    
    frase_calle = calle_norm
    candidatos_frase = []
    for name, name_norm, addr_norm, prov_norm_idx in candidatos_validos:
        combinado = f"{name_norm} {addr_norm}".strip()
        if frase_calle in combinado:
            candidatos_frase.append((name, name_norm, addr_norm))

    if len(candidatos_frase) == 1:
        return candidatos_frase[0][0], []
    elif len(candidatos_frase) > 1:
        candidatos_base = candidatos_frase
    else:
        # Strict Mode by user request: If phrase matching fails, fallback to NOTHING (Manual).
        candidatos_base = []

    loc_text = " ".join(str(x) for x in [localidad, ciudad, provincia] if isinstance(x, str))
    loc_norm = normalizar_texto(loc_text)
    loc_tokens = [t for t in loc_norm.split() if len(t) > 3]

    candidatos = candidatos_base
    if loc_tokens:
        candidatos_loc = []
        for name, name_norm, addr_norm in candidatos:
            combined = f"{name_norm} {addr_norm}"
            if any(tok in combined for tok in loc_tokens):
                candidatos_loc.append((name, name_norm, addr_norm))
        if len(candidatos_loc) == 1:
            return candidatos_loc[0][0], []
        elif len(candidatos_loc) > 1:
            candidatos = candidatos_loc

    num_digits = "".join(ch for ch in str(numero) if ch.isdigit())
    if not num_digits:
        return "", []

    refinados = []
    for name, name_norm, addr_norm in candidatos:
        combinado = f"{name_norm} {addr_norm}"
        nums_conf = re.findall(r"\b\d+\b", combinado)
        for t in nums_conf:
            if num_digits == t or num_digits.startswith(t) or t.startswith(num_digits):
                refinados.append(name)
                break

    if len(refinados) == 1:
        return refinados[0], []
    
    if refinados:
        return "", refinados
    
    # If we have candidates but couldn't refine them to 1 using the number,
    # DO NOT AUTO SELECT based on just the candidates list size being small or whatever.
    # Return as suggestions only.
    if candidatos:
        sugs = [c[0] for c in candidatos]
        return "", sugs[:10]

    return "", []

def extraer_base_localidad(valor):
    if not isinstance(valor, str) or not valor.strip():
        return ""
    norm = normalizar_texto(valor)
    tokens = norm.split()
    stop_tokens = {"CAPITAL", "CENTRO", "CIUDAD", "BARRIO", "NOROESTE", "NORESTE", "SUDOESTE", "SUDESTE", "NORTE", "SUR", "ESTE", "OESTE", "N", "S", "E", "O"}
    base_tokens = [t for t in tokens if t not in stop_tokens]
    return " ".join(base_tokens).strip()

def buscar_localidad_para_envio(indice_localidades, provincia, localidad, ciudad, cp):
    candidatos = indice_localidades
    cp_digits = ""
    if pd.notna(cp):
        s = str(cp)
        if s.endswith(".0"):s = s[:-2]
        cp_digits = "".join(ch for ch in s if ch.isdigit())

    if cp_digits:
        c_cp = [item for item in indice_localidades if item[4] == cp_digits]
        if len(c_cp) == 1:
            return c_cp[0][0], []
        elif len(c_cp) > 1:
            candidatos = c_cp

    loc_text = " ".join(str(x) for x in [localidad, ciudad] if isinstance(x, str))
    loc_norm = normalizar_texto(loc_text) if loc_text else ""
    loc_tokens = [t for t in loc_norm.split() if len(t) > 3]

    if loc_tokens and candidatos:
        c_loc = []
        for orig, norm, prov_norm_idx, loc_norm_idx, cp_idx in candidatos:
            combined = " ".join([norm, prov_norm_idx, loc_norm_idx])
            if any(tok in combined for tok in loc_tokens):
                c_loc.append((orig, norm, prov_norm_idx, loc_norm_idx, cp_idx))
        if len(c_loc) == 1:
            return c_loc[0][0], []
        elif len(c_loc) > 1:
            candidatos = c_loc

    loc_base_localidad = extraer_base_localidad(localidad)
    loc_base_ciudad    = extraer_base_localidad(ciudad)

    if (loc_base_localidad or loc_base_ciudad) and candidatos:
        exact = []
        for orig, norm, prov_norm_idx, loc_norm_idx, cp_idx in candidatos:
            if (loc_base_localidad and loc_norm_idx == loc_base_localidad) or (loc_base_ciudad and loc_norm_idx == loc_base_ciudad):
                exact.append((orig, norm, prov_norm_idx, loc_norm_idx, cp_idx))
        if len(exact) == 1:
            return exact[0][0], []
        elif len(exact) > 1:
            candidatos = exact

    prov_norm_q = ""
    if isinstance(provincia, str) and provincia.strip():
        prov_norm_q = normalizar_texto(provincia)
    if prov_norm_q and candidatos:
        c_prov = []
        for orig, norm, prov_norm_idx, loc_norm_idx, cp_idx in candidatos:
            if prov_norm_q in prov_norm_idx or prov_norm_q in norm:
                c_prov.append((orig, norm, prov_norm_idx, loc_norm_idx, cp_idx))
        if len(c_prov) == 1:
            return c_prov[0][0], []
        elif len(c_prov) > 1:
            candidatos = c_prov

    if prov_norm_q.find("CAPITAL FEDERAL") != -1 and candidatos:
        prioridad = []
        for orig, norm, prov_norm_idx, loc_norm_idx, cp_idx in candidatos:
            if "CIUDAD AUTONOMA BUENOS AIRES" in loc_norm_idx:
                prioridad.append((orig, norm, prov_norm_idx, loc_norm_idx, cp_idx))
        if prioridad:
            return prioridad[0][0], []

    if len(candidatos) == 1:
        return candidatos[0][0], []
    
    # Retornar sugerencias si hay candidatos que no filtraron a 1
    if candidatos:
        # candidatos es lista de tuplas
        sugs = [c[0] for c in candidatos]
        return "", sugs[:5]

    return "", []

# ========= API LOGIC =========

class AndreaniProcessor:
    def __init__(self, plantilla_path):
        self.plantilla_path = plantilla_path
        self.wb = load_workbook(plantilla_path, data_only=False)
        self.ws_conf = self.wb[HOJA_CONFIG]
        self.indice_sucursales, self.nombres_sucursales = construir_indice_sucursales(self.ws_conf)
        self.indice_localidades, self.nombres_localidades = construir_indice_localidades(self.ws_conf)
        # No guardamos cambios en self.wb aún, solo leemos config

    def process_csv(self, csv_content: bytes):
        ventas = pd.read_csv(io.BytesIO(csv_content), encoding="latin1", sep=";")
        
        if "Estado del envío" not in ventas.columns:
            return {"error": "No encuentro la columna 'Estado del envío'"}
            
        ventas_filtrado = ventas[ventas["Estado del envío"] == "Listo para enviar"].copy()
        
        if "Medio de envío" not in ventas_filtrado.columns:
            return {"error": "No columna Medio de envío"}
            
        if "Número de orden" not in ventas_filtrado.columns:
            return {"error": "No columna Número de orden"}

        ventas_filtrado = ventas_filtrado.sort_values("Número de orden").drop_duplicates("Número de orden", keep="first")
        
        # Sort so DOMICILIO comes first (D before S)
        # We can't sort dataframe effectively if we want strict control, better to sort the list of records later? 
        # Actually sorting the DF by "Número de orden" is good for ID order, but user wants Domicilio first.
        # We will sort the 'records' list at the end.
        
        records = []
        
        for _, r in ventas_filtrado.iterrows():
            item = {}
            medio = str(r["Medio de envío"] or "")
            is_sucursal = "Punto de retiro" in medio
            
            # Datos comunes
            nombre_envio = r.get("Nombre para el envío")
            if not isinstance(nombre_envio, str) or not nombre_envio.strip():
                nombre_envio = r.get("Nombre del comprador", "")
            nombre, apellido = split_nombre_apellido(nombre_envio)
            
            tel_envio = r.get("Teléfono para el envío")
            if (not isinstance(tel_envio, str) or not tel_envio.strip() or "no informado" in tel_envio.lower()):
                tel_envio = str(r.get("Teléfono", ""))
            cod_cel, num_cel = limpiar_telefono(tel_envio)
            
            item["nro_orden"] = str(r["Número de orden"])
            item["nombre"] = sanitizar_texto(nombre)
            item["apellido"] = sanitizar_texto(apellido)
            item["dni"] = sanitizar_texto(formatear_id(r.get("DNI / CUIT", "")))
            item["email"] = sanitizar_texto(r.get("Email", ""))
            item["cod_cel"] = cod_cel
            item["num_cel"] = num_cel
            item["calle"] = sanitizar_texto(r.get("Dirección", ""))
            item["numero"] = sanitizar_texto(limpiar_numero_calle(r.get("Número", "")))
            item["piso"] = sanitizar_piso(r.get("Piso", ""))
            
            obs = sanitizar_texto(r.get("Notas del comprador", ""))
            if len(obs) > 150:
                obs = obs[:150]
            item["observaciones"] = obs
            
            # Datos de direccion
            provincia = r.get("Provincia o estado", "")
            localidad = r.get("Localidad", "")
            ciudad    = r.get("Ciudad", "")
            cp        = r.get("Código postal", "")
            
            item["raw_provincia"] = sanitizar_texto(provincia)
            item["raw_localidad"] = sanitizar_texto(localidad)
            item["raw_ciudad"] = sanitizar_texto(ciudad)
            item["raw_cp"] = sanitizar_texto(cp)
            
            item["provincia_norm"] = normalizar_texto(provincia)
            
            item["tipo_envio"] = "SUCURSAL" if is_sucursal else "DOMICILIO"
            
            if is_sucursal:
                match, suggestions = buscar_sucursal_por_direccion(
                    self.indice_sucursales,
                    item["calle"], item["numero"], localidad, ciudad, provincia
                )
                item["match_value"] = match
                item["suggestions"] = suggestions
                item["status"] = "OK" if match else "MISSING"
            else:
                match, suggestions = buscar_localidad_para_envio(
                    self.indice_localidades,
                    provincia, localidad, ciudad, cp
                )
                item["match_value"] = match
                item["suggestions"] = suggestions
                item["status"] = "OK" if match else "MISSING"
            
            records.append(item)
            
        # User request: Sort by Type (DOMICILIO first) then by Order Number
        # "DOMICILIO" < "SUCURSAL", so standard sort works for type.
        # sort key: tuple(type, order_id_int)
        records.sort(key=lambda x: (x["tipo_envio"], int(x["nro_orden"]) if x["nro_orden"].isdigit() else x["nro_orden"]))

        return {
            "records": records,
            "meta": {
                "sucursales": self.nombres_sucursales,
                "localidades": self.nombres_localidades
            },
            "summary": {
                "total": len(records),
                "sucursal": sum(1 for r in records if r["tipo_envio"] == "SUCURSAL"),
                "domicilio": sum(1 for r in records if r["tipo_envio"] == "DOMICILIO"),
                "revisar": sum(1 for r in records if r["status"] == "MISSING")
            }
        }

    def generate_excel(self, verified_data, output_path):
        # Recargar plantilla limpia
        wb = load_workbook(self.plantilla_path, data_only=False)
        ws_dom = wb[HOJA_DOMICILIO]
        ws_suc = wb[HOJA_SUCURSAL]
        
        # Limpiar
        for ws in [ws_dom, ws_suc]:
            for row in range(FILA_INICIO, ULTIMA_FILA + 1):
                for col in range(1, 19 + 1):
                    ws.cell(row=row, column=col).value = None

        r_dom = FILA_INICIO
        r_suc = FILA_INICIO
        
        for item in verified_data:
            is_sucursal = item["tipo_envio"] == "SUCURSAL"
            ws = ws_suc if is_sucursal else ws_dom
            row = r_suc if is_sucursal else r_dom
            
            # Mapeo de columnas (1-based)
            # 1: Paquete (None)
            ws.cell(row=row, column=2).value = PESO_POR_DEFECTO_GR
            ws.cell(row=row, column=3).value = ALTO_DEF
            ws.cell(row=row, column=4).value = ANCHO_DEF
            ws.cell(row=row, column=5).value = PROF_DEF
            ws.cell(row=row, column=6).value = VALOR_DECLARADO
            ws.cell(row=row, column=7).value = str(item["nro_orden"])
            ws.cell(row=row, column=8).value = item["nombre"]
            ws.cell(row=row, column=9).value = item["apellido"]
            ws.cell(row=row, column=10).value = item["dni"]
            ws.cell(row=row, column=11).value = item["email"]
            ws.cell(row=row, column=12).value = item["cod_cel"]
            ws.cell(row=row, column=13).value = item["num_cel"]
            
            if is_sucursal:
                ws.cell(row=row, column=14).value = item["match_value"]
            else:
                ws.cell(row=row, column=14).value = item["calle"]
                ws.cell(row=row, column=15).value = item["numero"]
                ws.cell(row=row, column=16).value = item["piso"]
                ws.cell(row=row, column=17).value = "" # Dpto
                ws.cell(row=row, column=18).value = item["match_value"] 
                ws.cell(row=row, column=19).value = item["observaciones"]
            
            if is_sucursal:
                r_suc += 1
            else:
                r_dom += 1
                
        wb.save(output_path)
        return output_path
